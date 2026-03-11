"""Tests for citation matrix Excel export."""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from agent_harness.export.citation_matrix_excel import export_citation_matrix


@pytest.fixture
def sample_structured_result() -> dict:
    return {
        "operativo_id": "dce-test-001",
        "domain": "dce",
        "corrected_citation_matrix": [
            {
                "citation_text": "16 CFR Part 1303 — Total Lead in Surface Coatings",
                "original_verdict": "PASS",
                "corrected_verdict": "VALID",
                "corrected_rationale": "Valid citation, confirmed.",
                "correction_type": "confirmed",
            },
            {
                "citation_text": "compliance Section 101 — Lead Substrates",
                "original_verdict": "FAIL",
                "corrected_verdict": "INVALID",
                "corrected_rationale": "Must use 15 U.S.C. § 1278a format.",
                "correction_type": "confirmed",
            },
            {
                "citation_text": "SOR-2018-83 — CCPSA Total Lead (Canadian)",
                "original_verdict": "FAIL",
                "corrected_verdict": "INVALID",
                "corrected_rationale": "Canadian reg, no standing on US DCE.",
                "correction_type": "confirmed",
            },
            {
                "citation_text": "MISSING REQUIRED: 16 CFR § 1107.21 (Periodic Testing)",
                "original_verdict": "MISSING",
                "corrected_verdict": "NOT_APPLICABLE",
                "corrected_rationale": "Procedural requirement, not DCE-citable.",
                "correction_type": "verdict_fix",
            },
        ],
        "qa_summary": {
            "total_checks": 10,
            "blocking": 3,
            "warnings": 2,
            "info": 5,
        },
    }


def test_export_creates_xlsx(sample_structured_result: dict):
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "matrix.xlsx"
        export_citation_matrix(
            structured_result=sample_structured_result,
            pdf_filename="test-dce.pdf",
            output_path=path,
        )
        assert path.exists()
        wb = openpyxl.load_workbook(path)
        assert "Citation Matrix" in wb.sheetnames


def test_export_has_correct_columns(sample_structured_result: dict):
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "matrix.xlsx"
        export_citation_matrix(
            structured_result=sample_structured_result,
            pdf_filename="test-dce.pdf",
            output_path=path,
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Citation Matrix"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        # Original columns preserved + Santos review columns added
        assert "Filename" in headers
        assert "Citation" in headers
        assert "Source" in headers
        assert "Found in DCE" in headers
        assert "Applicability" in headers
        assert "Rationale" in headers
        assert "Santos Verdict" in headers
        assert "Correction Type" in headers
        assert "Santos Rationale" in headers


def test_export_has_correct_row_count(sample_structured_result: dict):
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "matrix.xlsx"
        export_citation_matrix(
            structured_result=sample_structured_result,
            pdf_filename="test-dce.pdf",
            output_path=path,
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Citation Matrix"]
        # Count citation data rows: rows where Filename (col 1) equals pdf_filename
        data_rows = [
            r for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value == "test-dce.pdf"
        ]
        assert len(data_rows) == 4


def test_export_maps_verdicts_to_applicability(sample_structured_result: dict):
    """Santos corrected_verdict maps to Applicability column."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "matrix.xlsx"
        export_citation_matrix(
            structured_result=sample_structured_result,
            pdf_filename="test-dce.pdf",
            output_path=path,
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Citation Matrix"]
        # Row 2: VALID -> applicable
        assert ws.cell(row=2, column=5).value == "applicable"
        # Row 3: INVALID -> not_applicable
        assert ws.cell(row=3, column=5).value == "not_applicable"
        # Row 5: NOT_APPLICABLE -> not_applicable
        assert ws.cell(row=5, column=5).value == "not_applicable"


def test_export_source_column_logic(sample_structured_result: dict):
    """Source column: 'dce' if found, 'robot' if MISSING, 'both' if original_verdict was PASS."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "matrix.xlsx"
        export_citation_matrix(
            structured_result=sample_structured_result,
            pdf_filename="test-dce.pdf",
            output_path=path,
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Citation Matrix"]
        # Row 2: PASS original -> both
        assert ws.cell(row=2, column=3).value == "both"
        # Row 3: FAIL original -> dce
        assert ws.cell(row=3, column=3).value == "dce"
        # Row 5: MISSING original -> robot
        assert ws.cell(row=5, column=3).value == "robot"


def test_export_summary_section(sample_structured_result: dict):
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "matrix.xlsx"
        export_citation_matrix(
            structured_result=sample_structured_result,
            pdf_filename="test-dce.pdf",
            output_path=path,
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Citation Matrix"]
        # Find summary rows
        values = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)
        }
        assert values.get("Total Citations") == 4


def test_export_empty_matrix():
    """Empty corrected_citation_matrix should produce a valid Excel with header + summary only."""
    result = {
        "operativo_id": "dce-empty",
        "corrected_citation_matrix": [],
        "qa_summary": {"total_checks": 0, "blocking": 0, "warnings": 0, "info": 0},
    }
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "empty.xlsx"
        export_citation_matrix(
            structured_result=result,
            pdf_filename="empty.pdf",
            output_path=path,
        )
        assert path.exists()
        wb = openpyxl.load_workbook(path)
        ws = wb["Citation Matrix"]
        # Header row exists
        assert ws.cell(row=1, column=1).value == "Filename"
        # No data rows (row 2 should be blank or summary)
        data_rows = [
            r for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value == "empty.pdf"
        ]
        assert len(data_rows) == 0


def test_export_unknown_verdict_falls_back_to_not_enough_info():
    """Unknown corrected_verdict should map to 'not_enough_info'."""
    result = {
        "operativo_id": "dce-unknown",
        "corrected_citation_matrix": [
            {
                "citation_text": "Some Citation",
                "original_verdict": "UNCERTAIN",
                "corrected_verdict": "UNKNOWN_VALUE",
                "corrected_rationale": "Unclear.",
                "correction_type": "confirmed",
            },
        ],
        "qa_summary": {},
    }
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "unknown.xlsx"
        export_citation_matrix(
            structured_result=result,
            pdf_filename="test.pdf",
            output_path=path,
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Citation Matrix"]
        assert ws.cell(row=2, column=5).value == "not_enough_info"


def test_export_rationale_column_shows_original_verdict(sample_structured_result: dict):
    """Column F (Rationale) shows original robot verdict, Column I shows Santos rationale."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "matrix.xlsx"
        export_citation_matrix(
            structured_result=sample_structured_result,
            pdf_filename="test-dce.pdf",
            output_path=path,
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Citation Matrix"]
        # Column F: original verdict reference
        assert "PASS" in ws.cell(row=2, column=6).value
        # Column I: Santos corrected rationale
        assert ws.cell(row=2, column=9).value == "Valid citation, confirmed."
