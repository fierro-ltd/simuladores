"""Export corrected citation matrix as Excel matching DCE Backend output format.

Takes the structured_result JSON from a completed DCE operativo and produces
an .xlsx file with the same columns as the DCE Backend's citation-matrix output,
plus Santos-QA review columns.

Existing columns (preserved):
    Filename | Citation | Source | Found in DCE | Applicability | Rationale

Added columns:
    Santos Verdict | Correction Type | Santos Rationale
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# -- Styles ------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_WRAP = Alignment(wrap_text=True, vertical="top")

_VERDICT_FILLS: dict[str, PatternFill] = {
    "VALID": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "INVALID": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "NOT_APPLICABLE": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "OVERREACH": PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"),
}

_APPLICABILITY_FILLS: dict[str, PatternFill] = {
    "applicable": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "not_applicable": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "not_enough_info": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}

_CORRECTION_FONT = Font(bold=True, color="C00000")

# Column widths matching the existing DCE Backend output
_COL_WIDTHS = {
    "A": 44,  # Filename
    "B": 55,  # Citation
    "C": 10,  # Source
    "D": 14,  # Found in DCE
    "E": 18,  # Applicability
    "F": 80,  # Rationale
    "G": 18,  # Santos Verdict
    "H": 20,  # Correction Type
    "I": 80,  # Santos Rationale
}


# -- Mapping helpers ----------------------------------------------------------

_VERDICT_TO_APPLICABILITY: dict[str, str] = {
    "VALID": "applicable",
    "INVALID": "not_applicable",
    "NOT_APPLICABLE": "not_applicable",
    "OVERREACH": "not_applicable",
}


def _infer_source(entry: dict[str, Any]) -> str:
    """Infer Source column from original_verdict."""
    ov = str(entry.get("original_verdict", "")).upper()
    if ov == "MISSING":
        return "robot"
    if ov == "PASS":
        return "both"
    return "dce"


def _infer_found_in_cpc(entry: dict[str, Any]) -> str:
    """Infer Found in DCE from original_verdict."""
    return "No" if str(entry.get("original_verdict", "")).upper() == "MISSING" else "Yes"


# -- Core export --------------------------------------------------------------


def export_citation_matrix(
    *,
    structured_result: dict[str, Any],
    pdf_filename: str,
    output_path: Path | str,
) -> Path:
    """Export corrected citation matrix to Excel.

    Args:
        structured_result: Parsed structured_result dict from the harness
            (must contain ``corrected_citation_matrix`` list).
        pdf_filename: Original PDF filename (for the Filename column).
        output_path: Where to write the .xlsx file.

    Returns:
        The output path (for chaining).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    matrix: list[dict[str, Any]] = structured_result.get("corrected_citation_matrix", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Citation Matrix"

    # -- Header ---------------------------------------------------------------
    headers = [
        "Filename", "Citation", "Source", "Found in DCE",
        "Applicability", "Rationale",
        "Santos Verdict", "Correction Type", "Santos Rationale",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = _THIN_BORDER

    # -- Data rows ------------------------------------------------------------
    for i, entry in enumerate(matrix, start=1):
        row = i + 1
        corrected_verdict = str(entry.get("corrected_verdict", ""))
        applicability = _VERDICT_TO_APPLICABILITY.get(corrected_verdict, "not_enough_info")
        correction_type = str(entry.get("correction_type", ""))

        ws.cell(row=row, column=1, value=pdf_filename).border = _THIN_BORDER
        ws.cell(row=row, column=2, value=entry.get("citation_text", "")).border = _THIN_BORDER

        ws.cell(row=row, column=3, value=_infer_source(entry)).border = _THIN_BORDER
        ws.cell(row=row, column=4, value=_infer_found_in_cpc(entry)).border = _THIN_BORDER

        cell_app = ws.cell(row=row, column=5, value=applicability)
        cell_app.border = _THIN_BORDER
        if applicability in _APPLICABILITY_FILLS:
            cell_app.fill = _APPLICABILITY_FILLS[applicability]

        # Rationale: brief original verdict note (DCE Backend column)
        original_verdict = str(entry.get("original_verdict", ""))
        ws.cell(row=row, column=6, value=f"Original Robot verdict: {original_verdict}").border = _THIN_BORDER
        ws.cell(row=row, column=6).alignment = _WRAP

        # Santos review columns
        cell_sv = ws.cell(row=row, column=7, value=corrected_verdict)
        cell_sv.border = _THIN_BORDER
        if corrected_verdict in _VERDICT_FILLS:
            cell_sv.fill = _VERDICT_FILLS[corrected_verdict]
            cell_sv.font = Font(bold=True)

        cell_ct = ws.cell(row=row, column=8, value=correction_type)
        cell_ct.border = _THIN_BORDER
        if correction_type in ("rationale_fix", "verdict_fix", "overreach_removed"):
            cell_ct.font = _CORRECTION_FONT

        ws.cell(row=row, column=9, value=entry.get("corrected_rationale", "")).border = _THIN_BORDER
        ws.cell(row=row, column=9).alignment = _WRAP

    # -- Summary section ------------------------------------------------------
    data_count = len(matrix)
    summary_start = data_count + 3  # blank row after data

    ws.cell(row=summary_start, column=1, value="Summary").font = Font(bold=True, size=12)

    qa = structured_result.get("qa_summary", {})
    if isinstance(qa, dict):
        summary_text = (
            f"Santos-QA reviewed {data_count} citations. "
            f"QA checks: {qa.get('total_checks', 0)} total, "
            f"{qa.get('blocking', 0)} blocking, "
            f"{qa.get('warnings', 0)} warnings."
        )
    else:
        summary_text = f"Santos-QA reviewed {data_count} citations."
    ws.cell(row=summary_start + 1, column=1, value=summary_text)

    # Counts
    counts_start = summary_start + 3
    valid = sum(1 for e in matrix if e.get("corrected_verdict") == "VALID")
    invalid = sum(1 for e in matrix if e.get("corrected_verdict") == "INVALID")
    na = sum(1 for e in matrix if e.get("corrected_verdict") == "NOT_APPLICABLE")
    overreach = sum(1 for e in matrix if e.get("corrected_verdict") == "OVERREACH")
    confirmed = sum(1 for e in matrix if e.get("correction_type") == "confirmed")
    changed = data_count - confirmed

    stats = [
        ("Total Citations", data_count),
        ("Total Applicable", valid),
        ("Total Not Applicable", invalid + na + overreach),
        ("Total Not Enough Info", 0),
        ("", ""),
        ("Verdicts Changed by Santos-QA", changed),
        ("Verdicts Confirmed", confirmed),
    ]
    for j, (label, val) in enumerate(stats):
        ws.cell(row=counts_start + j, column=1, value=label).border = _THIN_BORDER
        ws.cell(row=counts_start + j, column=2, value=val).border = _THIN_BORDER

    # -- Column widths --------------------------------------------------------
    for letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    wb.save(output_path)
    return output_path
